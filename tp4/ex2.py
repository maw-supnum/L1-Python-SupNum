# Exercice 2: Traitement des données COVID-19
# 1. Créer la feuille "Bilan"
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook


def creer_feuille_bilan(fichier_excel):
    try:
        # Charger le classeur existant
        workbook = load_workbook(fichier_excel)

        # Vérifier si la feuille existe déjà
        if "Bilan" in workbook.sheetnames:
            # Utiliser la feuille existante
            sheet = workbook["Bilan"]
        else:
            # Créer une nouvelle feuille
            sheet = workbook.create_sheet("Bilan")

        # Configurer les en-têtes
        sheet["B1"] = ""
        sheet["C1"] = ""
        sheet["B2"] = ""
        sheet["C2"] = "Mauritanie"
        sheet["B3"] = "1- Nombre total de cas"
        sheet["B4"] = "2- Nombre de cas pic"
        sheet["B5"] = "3- Date du jour pic en nombre de cas"

        # Enregistrer les modifications
        workbook.save(fichier_excel)

        print(f"Feuille 'Bilan' créée dans le fichier '{fichier_excel}'.")

        return True
    except Exception as e:
        print(f"Erreur lors de la création de la feuille 'Bilan': {e}")
        return False


# 2. Fonction Effacer()
def Effacer(fichier_excel):
    try:
        # Charger le classeur
        workbook = load_workbook(fichier_excel)

        # Vérifier si la feuille "Bilan" existe
        if "Bilan" not in workbook.sheetnames:
            print("Erreur: La feuille 'Bilan' n'existe pas.")
            return False

        # Accéder à la feuille
        sheet = workbook["Bilan"]

        # Effacer les cellules C3, C4 et C5
        sheet["C3"] = None
        sheet["C4"] = None
        sheet["C5"] = None

        # Enregistrer les modifications
        workbook.save(fichier_excel)

        print("Les cellules C3, C4 et C5 ont été effacées.")

        return True
    except Exception as e:
        print(f"Erreur lors de l'effacement des cellules: {e}")
        return False


# 3. Fonction Calculer()
def Calculer(fichier_excel):
    try:
        # Lire les données de la feuille "Data"
        df = pd.read_excel(fichier_excel, sheet_name="Data")

        # Calculer le nombre total de cas
        nombre_total_cas = df["Nombre de cas"].sum()

        # Trouver le pic (nombre maximal de cas)
        nombre_cas_pic = df["Nombre de cas"].max()

        # Trouver la date du pic
        index_pic = df["Nombre de cas"].idxmax()
        date_pic = df.loc[index_pic, "Date"]

        # Si la date est un timestamp ou datetime, la formater
        if isinstance(date_pic, (pd.Timestamp, datetime)):
            date_pic = date_pic.strftime("%d/%m/%Y")

        # Charger le classeur pour mettre à jour les cellules
        workbook = load_workbook(fichier_excel)
        sheet = workbook["Bilan"]

        # Mettre à jour les cellules
        sheet["C3"] = nombre_total_cas
        sheet["C4"] = nombre_cas_pic
        sheet["C5"] = date_pic

        # Enregistrer les modifications
        workbook.save(fichier_excel)

        print("Calculs effectués et cellules mises à jour:")
        print(f"1- Nombre total de cas: {nombre_total_cas}")
        print(f"2- Nombre de cas pic: {nombre_cas_pic}")
        print(f"3- Date du jour pic: {date_pic}")

        return True
    except Exception as e:
        print(f"Erreur lors du calcul et de la mise à jour des cellules: {e}")
        return False
