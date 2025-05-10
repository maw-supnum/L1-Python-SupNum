# TP4 - Traitement de fichiers Excel
# Institut Supérieur du Numérique (SupNum)

import pandas as pd
from openpyxl import load_workbook


# Exercice 1 : Traitement des notes
# Q1: Afficher les notes d'un étudiant à partir de son matricule
def afficher_notes_etudiant(fichier_excel, matricule):
    try:
        # Lire le fichier Excel
        df = pd.read_excel(fichier_excel)

        # Chercher l'étudiant par matricule
        etudiant = df[df["Matricule"] == matricule]

        if etudiant.empty:
            print(
                f"Erreur: Aucun étudiant avec le matricule {matricule} n'a été trouvé."
            )
            return False

        # Afficher les notes
        note_ds = etudiant["Note DS"].values[0]
        note_examen = etudiant["Note Examen"].values[0]

        print(f"Notes de l'étudiant avec matricule {matricule}:")
        print(f"Note DS: {note_ds}")
        print(f"Note Examen: {note_examen}")

        return True
    except Exception as e:
        print(f"Erreur lors de la lecture du fichier Excel: {e}")
        return False


# Q2: Afficher le matricule de l'étudiant ayant la meilleure note d'examen
def meilleure_note_examen(fichier_excel):
    try:
        # Lire le fichier Excel
        df = pd.read_excel(fichier_excel)

        # Trouver l'index de la note d'examen maximale
        index_max = df["Note Examen"].idxmax()

        # Obtenir le matricule et la note correspondante
        matricule = df.loc[index_max, "Matricule"]
        note_max = df.loc[index_max, "Note Examen"]

        print(
            f"L'étudiant avec le matricule {matricule} a la meilleure note d'examen: {note_max}"
        )

        return matricule, note_max
    except Exception as e:
        print(f"Erreur lors de la recherche de la meilleure note: {e}")
        return None, None


# Q3: Créer un nouveau fichier Excel avec les moyennes
def creer_fichier_moyennes(fichier_source, fichier_destination):
    try:
        # Lire le fichier Excel source
        df = pd.read_excel(fichier_source)

        # Calculer les moyennes (40% DS + 60% Examen)
        df["Moyenne"] = 0.4 * df["Note DS"] + 0.6 * df["Note Examen"]

        # Arrondir à 2 décimales
        df["Moyenne"] = df["Moyenne"].round(2)

        # Créer un nouveau DataFrame avec seulement les colonnes nécessaires
        df_moyennes = df[["Matricule", "Moyenne"]]

        # Enregistrer dans un nouveau fichier Excel
        df_moyennes.to_excel(fichier_destination, index=False)

        print(f"Fichier '{fichier_destination}' créé avec succès.")

        return True
    except Exception as e:
        print(f"Erreur lors de la création du fichier des moyennes: {e}")
        return False


# Q4: Ajouter une feuille "Moyenne" au fichier Excel existant
def ajouter_feuille_moyenne(fichier_excel):
    try:
        # Lire les données avec pandas
        df = pd.read_excel(fichier_excel)

        # Calculer les moyennes (40% DS + 60% Examen)
        df["Moyenne"] = 0.4 * df["Note DS"] + 0.6 * df["Note Examen"]

        # Arrondir à 2 décimales
        df["Moyenne"] = df["Moyenne"].round(2)

        # Créer un nouveau DataFrame avec seulement les colonnes nécessaires
        df_moyennes = df[["Matricule", "Moyenne"]]

        # Charger le fichier Excel existant
        workbook = load_workbook(fichier_excel)

        # Vérifier si la feuille existe déjà
        if "Moyenne" in workbook.sheetnames:
            # Supprimer la feuille existante
            sheet = workbook["Moyenne"]
            workbook.remove(sheet)

        # Créer un writer ExcelWriter avec le mode d'ajout
        with pd.ExcelWriter(fichier_excel, engine="openpyxl", mode="a") as writer:
            writer.book = workbook
            df_moyennes.to_excel(writer, sheet_name="Moyenne", index=False)

        print(f"Feuille 'Moyenne' ajoutée au fichier '{fichier_excel}' avec succès.")

        return True
    except Exception as e:
        print(f"Erreur lors de l'ajout de la feuille des moyennes: {e}")
        return False
