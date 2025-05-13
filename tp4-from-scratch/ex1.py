import openpyxl
from openpyxl import Workbook


def rechercher_notes_par_matricule():
    """
    Q1: Programme qui demande la saisie d'un matricule d'étudiant et affiche ses deux notes.
    Affiche un message d'erreur si le matricule n'apparaît pas dans le fichier.
    """
    # Charger le classeur Excel
    classeur = openpyxl.load_workbook("Notes_DEV110.xlsx")

    # Sélectionner la première feuille
    feuille = classeur.active

    # Demander le matricule à l'utilisateur
    matricule_recherche = int(input("Entrez le matricule de l'étudiant: "))

    # Rechercher le matricule dans le fichier
    matricule_trouve = False

    # Parcourir les lignes du fichier (on suppose que les données commencent à la ligne 2)
    for ligne in feuille.iter_rows(min_row=2, values_only=True):
        if ligne[0] == matricule_recherche:  # Première colonne contient le matricule
            print(f"Matricule: {ligne[0]}")
            print(f"Note DS: {ligne[1]}")
            print(f"Note Examen: {ligne[2]}")
            matricule_trouve = True
            break

    # Afficher un message d'erreur si le matricule n'est pas trouvé
    if not matricule_trouve:
        print(f"Erreur: Le matricule {matricule_recherche} n'existe pas dans le fichier.")

    # Fermer le classeur
    classeur.close()


def meilleure_note_examen():
    """
    Q2: Programme qui affiche le matricule de l'étudiant ayant la meilleure note d'examen.
    """
    # Charger le classeur Excel
    classeur = openpyxl.load_workbook("Notes_DEV110.xlsx")

    # Sélectionner la première feuille
    feuille = classeur.active

    meilleure_note = -1
    meilleur_matricule = ""

    # Parcourir les lignes du fichier (on suppose que les données commencent à la ligne 2)
    for ligne in feuille.iter_rows(min_row=2, values_only=True):
        matricule = ligne[0]
        note_examen = ligne[2]

        # Vérifier si cette note est la meilleure jusqu'à présent
        if note_examen and note_examen > meilleure_note:
            meilleure_note = note_examen
            meilleur_matricule = matricule

    print(f"Le matricule de l'étudiant ayant la meilleure note d'examen est: {meilleur_matricule}")
    print(f"Sa note d'examen est: {meilleure_note}")

    # Fermer le classeur
    classeur.close()


def creer_fichier_moyennes():
    """
    Q3: Programme qui crée un fichier Excel nommé Moyennes.xlsx contenant
    les matricules et les moyennes en algo et C++ (40% DS et 60% examen).
    """
    # Charger le classeur Excel source
    classeur_source = openpyxl.load_workbook("Notes_DEV110.xlsx")
    feuille_source = classeur_source.active

    # Créer un nouveau classeur pour les moyennes
    classeur_moyennes = Workbook()
    feuille_moyennes = classeur_moyennes.active

    # Ajouter les en-têtes
    feuille_moyennes['A1'] = "Matricule"
    feuille_moyennes['B1'] = "Moyenne"

    # Parcourir les données et calculer les moyennes
    ligne_dest = 2  # Commencer à écrire à la ligne 2

    for ligne in feuille_source.iter_rows(min_row=2, values_only=True):
        matricule = ligne[0]
        note_ds = ligne[1]
        note_examen = ligne[2]

        # Calculer la moyenne (40% DS + 60% examen)
        moyenne = ((note_ds if note_ds else 0) * 0.4) + ((note_examen if note_examen else 0) * 0.6)

        # Écrire les données dans le nouveau fichier
        feuille_moyennes.cell(row=ligne_dest, column=1, value=matricule)
        feuille_moyennes.cell(row=ligne_dest, column=2, value=round(moyenne, 2))

        ligne_dest += 1

    # Enregistrer le nouveau classeur
    classeur_moyennes.save("Moyennes.xlsx")
    print("Le fichier Moyennes.xlsx a été créé avec succès.")

    # Fermer les classeurs
    classeur_source.close()
    classeur_moyennes.close()


def ajouter_feuille_moyenne():
    """
    Q4: Programme qui crée une feuille Moyenne dans le classeur Excel Notes_DEV110.xlsx
    contenant les matricules et les moyennes.
    """
    # Charger le classeur Excel
    classeur = openpyxl.load_workbook("Notes_DEV110.xlsx")

    # Sélectionner la première feuille (pour lire les données)
    feuille_source = classeur.active

    # Créer une nouvelle feuille pour les moyennes (ou la supprimer si elle existe déjà)
    if "Moyenne" in classeur.sheetnames:
        classeur.remove(classeur["Moyenne"])

    feuille_moyenne = classeur.create_sheet("Moyenne")

    # Ajouter les en-têtes
    feuille_moyenne['A1'] = "Matricule"
    feuille_moyenne['B1'] = "Moyenne"

    # Parcourir les données et calculer les moyennes
    ligne_dest = 2  # Commencer à écrire à la ligne 2

    for ligne in feuille_source.iter_rows(min_row=2, values_only=True):
        matricule = ligne[0]
        note_ds = ligne[1]
        note_examen = ligne[2]

        # Calculer la moyenne (40% DS + 60% examen)
        moyenne = ((note_ds if note_ds else 0) * 0.4) + ((note_examen if note_examen else 0) * 0.6)

        # Écrire les données dans la nouvelle feuille
        feuille_moyenne.cell(row=ligne_dest, column=1, value=matricule)
        feuille_moyenne.cell(row=ligne_dest, column=2, value=round(moyenne, 2))

        ligne_dest += 1

    # Enregistrer le classeur
    classeur.save("Notes_DEV110.xlsx")
    print("La feuille 'Moyenne' a été ajoutée au classeur Notes_DEV110.xlsx avec succès.")

    # Fermer le classeur
    classeur.close()


# Menu principal pour tester les fonctions
if __name__ == "__main__":
    print("\nTP N°4 - Exercice 1: Traitement de fichiers Excel\n")

    choix = input("\nEntrez votre choix (1-4): ")

    if choix == "1":
        print("1. Rechercher les notes d'un étudiant par matricule")
        rechercher_notes_par_matricule()
    elif choix == "2":
        print("2. Afficher le matricule de l'étudiant ayant la meilleure note d'examen")
        meilleure_note_examen()
    elif choix == "3":
        print("3. Créer un fichier Moyennes.xlsx")
        creer_fichier_moyennes()
    elif choix == "4":
        print("4. Ajouter une feuille Moyenne au classeur Notes_DEV110.xlsx")
        ajouter_feuille_moyenne()
    else:
        print("Choix invalide.")