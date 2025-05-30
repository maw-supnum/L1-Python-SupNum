import openpyxl
from openpyxl.styles import PatternFill


def creer_feuille_bilan():
    """
    1. Créer la feuille « Bilan » dans le classeur Covid.xlsx.
    """
    # Charger le classeur Excel
    classeur = openpyxl.load_workbook("Covid.xlsx")

    # Vérifier si la feuille Bilan existe déjà
    if "Bilan" in classeur.sheetnames:
        # Supprimer la feuille existante
        classeur.remove(classeur["Bilan"])

    # Créer la nouvelle feuille Bilan
    feuille_bilan = classeur.create_sheet("Bilan")

    # Remplir les en-têtes et les libellés comme dans l'exemple
    feuille_bilan['B3'] = "1- Nombre total de cas"
    feuille_bilan['B4'] = "2- Nombre de cas pic"
    feuille_bilan['B5'] = "3- Date du jour pic en nombre de cas"
    feuille_bilan['C2'] = "Mauritanie"

    fill = PatternFill(
        start_color="FFFF99",
        fill_type="solid"
    )

    feuille_bilan['B3'].fill = fill
    feuille_bilan['B4'].fill = fill
    feuille_bilan['B5'].fill = fill


    # Enregistrer le classeur
    classeur.save("Covid.xlsx")
    print("La feuille 'Bilan' a été créée avec succès.")

    # Fermer le classeur
    classeur.close()


def effacer():
    """
    Fonction qui vide le contenu des cellules C3, C4 et C5 de la feuille Bilan.
    """
    # Charger le classeur Excel
    classeur = openpyxl.load_workbook("Covid.xlsx")

    # Vérifier si la feuille Bilan existe
    if "Bilan" not in classeur.sheetnames:
        print("La feuille Bilan n'existe pas.")
        classeur.close()
        return

    # Sélectionner la feuille Bilan
    feuille_bilan = classeur["Bilan"]

    # Vider les cellules C3, C4 et C5
    feuille_bilan['C3'] = None
    feuille_bilan['C4'] = None
    feuille_bilan['C5'] = None

    # Enregistrer le classeur
    classeur.save("Covid.xlsx")
    print("Les cellules C3, C4 et C5 ont été vidées avec succès.")

    # Fermer le classeur
    classeur.close()


def calculer():
    """
    Fonction qui remplit automatiquement les trois cellules en effectuant un seul
    parcours de la feuille Data.
    """
    # Charger le classeur Excel
    classeur = openpyxl.load_workbook("tp4-from-scratch/Covid.xlsx")

    # Vérifier si les feuilles Bilan et Data existent
    if "Bilan" not in classeur.sheetnames:
        print("La feuille Bilan n'existe pas.")
        classeur.close()
        return False

    if "Data" not in classeur.sheetnames:
        print("La feuille Data n'existe pas.")
        classeur.close()
        return False

    # Sélectionner les feuilles Bilan et Data
    feuille_bilan = classeur["Bilan"]
    feuille_data = classeur["Data"]

    # Variables pour stocker les résultats
    nombre_total_cas = 0
    nombre_cas_pic = 0
    date_jour_pic = None

    # Parcourir la feuille Data en une seule fois
    # On suppose que les données commencent à la ligne 2, avec des en-têtes à la ligne 1
    # et que la colonne avec le nombre de cas est à l'index 1 (colonne B)
    # et que la colonne avec la date est à l'index 0 (colonne A)
    for ligne in feuille_data.iter_rows(min_row=2, values_only=True):
        if (ligne[2]).lower().strip() != "mauritania":
            continue
        date = ligne[-1]  # Date (colonne A)
        cas = ligne[0]  # Nombre de cas (colonne B)

        # Ignorer les lignes avec des valeurs manquantes
        if cas is None:
            continue

        # Mettre à jour le nombre total de cas
        nombre_total_cas += cas

        # Mettre à jour le nombre de cas pic et la date correspondante
        if cas > nombre_cas_pic:
            nombre_cas_pic = cas
            date_jour_pic = date

    # Remplir les cellules C3, C4 et C5 avec les résultats
    feuille_bilan['C3'] = nombre_total_cas
    feuille_bilan['C4'] = nombre_cas_pic
    feuille_bilan['C5'] = date_jour_pic

    # Enregistrer le classeur
    classeur.save("tp4-from-scratch/Covid.xlsx")
    print("Les calculs ont été effectués et les cellules ont été remplies avec succès.")

    # Fermer le classeur
    classeur.close()

    return None


# Menu principal pour tester les fonctions
if __name__ == "__main__":
    print("\nTP N°4 - Exercice 2: Traitement des données COVID-19\n")
    print("1. Créer la feuille Bilan")
    print("2. Effacer les cellules C3, C4 et C5")
    print("3. Calculer et remplir les cellules C3, C4 et C5")

    choix = input("\nEntrez votre choix (1-3): ")

    if choix == "1":
        creer_feuille_bilan()
    elif choix == "2":
        effacer()
    elif choix == "3":
        calculer()
    else:
        print("Choix invalide.")